from openpyxl import load_workbook
import requests, re
from bs4 import BeautifulSoup
from requests.auth import HTTPDigestAuth

YL_host = load_workbook(r'D:\Download\HostResourceExport-2019-12-23.xlsx')
YL_host_sheetnamelist = YL_host.sheetnames
YL_host_sheet = YL_host[YL_host_sheetnamelist[0]]
JY_host = load_workbook(r'D:\Download\HostResourceExport-2019-12-23 (1).xlsx')
JY_host_sheetnamelist = JY_host.sheetnames
JY_host_sheet = JY_host[JY_host_sheetnamelist[0]]
YL_vm = load_workbook(r'D:\Download\VmExport-2019-12-23.xlsx')
YL_vm_sheetnamelist = YL_vm.sheetnames
YL_vm_sheet = YL_vm[YL_vm_sheetnamelist[0]]
JY_vm = load_workbook(r'D:\Download\VmExport-2019-12-23 (1).xlsx')
JY_vm_sheetnamelist = JY_vm.sheetnames
JY_vm_sheet = JY_vm[JY_vm_sheetnamelist[0]]
finalResExcel = load_workbook(r'C:\Users\gaosi\Desktop\Work\资源统计\云平台资源信息-2019.12.16.xlsx')
finalResExcel_sheetlist = finalResExcel.sheetnames
FRE_Hostinfo = finalResExcel[finalResExcel_sheetlist[1]]


# 左右边一列 各个主机运行虚拟机数量
def getVmNum(sheet):
    list = []
    mkNum = 0
    for rowNum in range(1, sheet.max_row + 1):
        Value = sheet.cell(row=rowNum, column=1).value
        if Value == 'move-cvk':
            mkNum = rowNum
            break
    for rowNum in range(2, sheet.max_row + 1):
        Value = sheet.cell(row=rowNum, column=5).value
        endNum = Value.find('启动') - 1
        # 主机上存在的虚拟机数在‘启动’前两格 需要去掉‘[’ 并判断是否为move-cvk
        if rowNum != mkNum:
            list.append(int(Value[0:endNum]))
    return list


# cpu利用率、cpu分配比、内存利用率、内存分配比都可以使用这个函数处理
def getRate(sheet, colnumber):
    list = []
    mkNum = 0
    for rowNum in range(1, sheet.max_row + 1):
        Value = sheet.cell(row=rowNum, column=1).value
        if Value == 'move-cvk':
            mkNum = rowNum
            break
    for rowNum in range(2, sheet.max_row + 1):
        Value = sheet.cell(row=rowNum, column=colnumber).value
        # 表格中的数据后缀有一个 ‘%’ 将其舍去 并判断是不是move-cvk 如果是则不执行
        if rowNum != mkNum:
            list.append(float(Value[:-1]))
    return list


# 插入虚拟机数据
def copyVmInfo(dataSheet, targetSheet):
    number = 1
    for row in dataSheet.iter_rows(min_row=1, max_row=dataSheet.max_row, values_only=True):
        litter = 'A'
        for value in row:
            cell = litter + str(number)
            targetSheet[cell] = value
            litter = chr(ord(litter) + 1)
        number += 1


# cpu利用率
cpuRateYL = getRate(YL_host_sheet, 8)
cpuRateJY = getRate(JY_host_sheet, 8)

# 内存利用率
memRateYL = getRate(YL_host_sheet, 9)
memRateJY = getRate(JY_host_sheet, 9)

# cpu分配率
cpuDisRateYL = getRate(YL_host_sheet, 10)
cpuDisRateJY = getRate(JY_host_sheet, 10)

# 内存分配率
memDisRateYL = getRate(YL_host_sheet, 11)
memDisRateJY = getRate(JY_host_sheet, 11)

# 主机下虚拟机数量
vmNumYL = getVmNum(YL_host_sheet)
vmNumJY = getVmNum(JY_host_sheet)


# 向Excel中插入列表数据
def insertDatatoFRE(startRowNum, columnNum, dataListName):
    for i in dataListName:
        FRE_Hostinfo.cell(row=startRowNum, column=columnNum).value = i
        startRowNum += 1


# 插入主机数据
insertDatatoFRE(4, 6, cpuRateYL)
insertDatatoFRE(4, 7, memRateYL)
insertDatatoFRE(4, 12, cpuDisRateYL)
insertDatatoFRE(4, 13, memDisRateYL)
insertDatatoFRE(4, 15, vmNumYL)
insertDatatoFRE(16, 6, cpuRateJY)
insertDatatoFRE(16, 7, memRateJY)
insertDatatoFRE(16, 12, cpuDisRateJY)
insertDatatoFRE(16, 13, memDisRateJY)
insertDatatoFRE(16, 15, vmNumJY)

# 插入虚拟机信息
FRE_vmInfoYL = finalResExcel[finalResExcel_sheetlist[3]]
copyVmInfo(YL_vm_sheet, FRE_vmInfoYL)
FRE_vmInfoJY = finalResExcel[finalResExcel_sheetlist[4]]
copyVmInfo(JY_vm_sheet, FRE_vmInfoJY)


# finalResExcel.save('C:/Users/gaosi/Desktop/Work/资源统计/云平台资源信息test-2019.12.16.xlsx')
# print ('C:/Users/gaosi/Desktop/Work/资源统计/云平台资源信息test-2019.12.16.xlsx已生成。')

# 获取信息写入列表
def Addinfotolist(info_name, file_name):
    info = file_name.findAll(info_name)
    list = []
    for i in info:
        list.append(i.string)
    return list

totalSizeYL, freeSizeYL = [], []
cpuCountListYL, memSizeListYL = [], []
#通过接口获取医疗云主机的虚拟cpu总数，内存大小
def YLInfo():
    # 获取主机的id，根据主机的id查询对应主机的虚拟cpu数目以及内存大小
    hostId_url = "http://10.10.0.7:8080/cas/casrs/host/"
    resp = requests.get(hostId_url, auth=HTTPDigestAuth('yd_gaosi', 'jsdk@123'))
    hostIdhtml = resp.text
    soupHostIdYL = BeautifulSoup(hostIdhtml, "html.parser")
    host_id = Addinfotolist('id', soupHostIdYL)
    # 删除move-cvk的id
    del host_id[Addinfotolist('name', soupHostIdYL).index('move-cvk')]
    print(host_id)
    # 循环通过主机id查询对应主机的虚拟cpu数量以及内存大小
    for id in host_id:
        hostInfo_url = "http://10.10.0.7:8080/cas/casrs/host/id/%s" % id
        resp = requests.get(hostInfo_url, auth=HTTPDigestAuth('yd_gaosi', 'jsdk@123'))
        hostInfoHtml = resp.text
        hostInfoHtml = re.sub("cpuCount", "cpucount", hostInfoHtml, count=0, flags=0)
        hostInfoHtml = re.sub("memorySize", "memorysize", hostInfoHtml, count=0, flags=0)
        soupHostInfoYL = BeautifulSoup(hostInfoHtml, "html.parser")
        cpuCountListYL.append(int(soupHostInfoYL.find('cpucount').string))
        memSizeListYL.append(round(float(soupHostInfoYL.find('memorysize').string) / 1024, 2))

    # 调取接口查询ipsan存储总值和实际可用大小为大

    ShareFile_url = 'http://10.10.0.7:8080/cas/casrs/host/id/%s/storage' % host_id[0]
    resp = requests.get(ShareFile_url, auth=HTTPDigestAuth('yd_gaosi', 'jsdk@123'))
    shareFileInfo = resp.text
    shareFileInfo = re.sub('totalSize', 'totalsize', shareFileInfo, count=0, flags=0)
    shareFileInfo = re.sub('freeSize', 'freesize', shareFileInfo, count=0, flags=0)
    soupSFYL = BeautifulSoup(shareFileInfo, "html.parser")
    #将存储的单位换算成TB然后存入列表中
    totalSizeSoupYL = soupSFYL.findAll('totalsize')
    for i in totalSizeSoupYL:
        totalSizeYL.append(round(int(i.string) / 1024 ** 2, 2))
    freeSizeSoupYL = soupSFYL.findAll('freesize')
    for i in freeSizeSoupYL:
        freeSizeYL.append(round(int(i.string) / 1024 ** 2, 2))
    print(cpuCountListYL, '\n', memSizeListYL, '\n', totalSizeYL, '\n', freeSizeYL)
    return 0

cpuCountListJY, memSizeListJY = [], []
totalSizeJY, freeSizeJY = [], []
#通过接口获取医疗云主机的虚拟cpu总数，内存大小
def JYInfo():
    # 获取主机的id，根据主机的id查询对应主机的虚拟cpu数目以及内存大小
    hostId_url = "http://10.20.0.7:8080/cas/casrs/host/"
    resp = requests.get(hostId_url, auth=HTTPDigestAuth('yd_gaosi', 'jsdk@123'))
    hostIdhtml = resp.text
    soupHostIdJY = BeautifulSoup(hostIdhtml, "html.parser")
    host_id = Addinfotolist('id', soupHostIdJY)
    # 循环通过主机id查询对应主机的虚拟cpu数量以及内存大小
    for id in host_id:
        hostInfo_url = "http://10.20.0.7:8080/cas/casrs/host/id/%s" % id
        resp = requests.get(hostInfo_url, auth=HTTPDigestAuth('yd_gaosi', 'jsdk@123'))
        hostInfoHtml = resp.text
        hostInfoHtml = re.sub("cpuCount", "cpucount", hostInfoHtml, count=0, flags=0)
        hostInfoHtml = re.sub("memorySize", "memorysize", hostInfoHtml, count=0, flags=0)
        soupHostInfoJY = BeautifulSoup(hostInfoHtml, "html.parser")
        cpuCountListJY.append(int(soupHostInfoJY.find('cpucount').string))
        memSizeListJY.append(round(float(soupHostInfoJY.find('memorysize').string) / 1024, 2))

    # 调取接口查询ipsan存储总值和实际可用大小为大
    ShareFile_url = 'http://10.20.0.7:8080/cas/casrs/host/id/%s/storage' % host_id[0]
    resp = requests.get(ShareFile_url, auth=HTTPDigestAuth('yd_gaosi', 'jsdk@123'))
    shareFileInfo = resp.text
    shareFileInfo = re.sub('totalSize', 'totalsize', shareFileInfo, count=0, flags=0)
    shareFileInfo = re.sub('freeSize', 'freesize', shareFileInfo, count=0, flags=0)
    soupSFJY = BeautifulSoup(shareFileInfo, "html.parser")
    #将存储的单位换算成TB然后存入列表中
    totalSizeSoupJY = soupSFJY.findAll('totalsize')
    for i in totalSizeSoupJY:
        totalSizeJY.append(round(int(i.string) / 1024 ** 2, 2))
    freeSizeSoupJY = soupSFJY.findAll('freesize')
    for i in freeSizeSoupJY:
        freeSizeJY.append(round(int(i.string) / 1024 ** 2, 2))
    print(cpuCountListJY, '\n', memSizeListJY, '\n', totalSizeJY, '\n', freeSizeJY)
    return 0

YLInfo()
JYInfo()

insertDatatoFRE(3, 4, cpuCountListYL)
insertDatatoFRE(4, 4, memSizeListYL)
insertDatatoFRE(16, 3, cpuCountListJY)
insertDatatoFRE(16, 4, memSizeListJY)

cpuAllocatedResYL, memAllocatedResYL = [], []
cpuAllocatedResJY, memAllocatedResJY = [], []
print(cpuAllocatedResYL, memAllocatedResYL)

for num in range(1, len(cpuCountListYL)):
    cpuAllocatedResYL.append(num)

finalResExcel.save('C:/Users/gaosi/Desktop/Work/资源统计/云平台资源信息test-2019.12.23.xlsx')
print('C:/Users/gaosi/Desktop/Work/资源统计/云平台资源信息test-2019.12.23.xlsx已生成。')
