from openpyxl import load_workbook
import datetime
import re

GD_excel = load_workbook(r'C:\Users\gaosi\Desktop\Work\质检\故障工单查询导出(20191201-20191208).xlsx')
GD_sheet = GD_excel['故障工单查询导出']
# GD_Error = GD_excel.create_sheet('问题工单汇总')

# deprecationTime = []
#
# def troubleShootingTimeCompare(rownum):
#     cell1 = 'L' + str(rownum)
#     troubleShootingTimeCell = GD_sheet[cell1]
#     troubleShootingTimeRE = re.match(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}', troubleShootingTimeCell.value).group()
#     #print (troubleShootingTimeRE)
#     cell2 = 'R' + str(rownum)
#     descriptionCell = GD_sheet[cell2]
#     descriptionCellRE = re.findall(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}', descriptionCell.value)
#     #print (descriptionCellRE[0])
#     if troubleShootingTimeRE == descriptionCellRE[0] :
#         deprecationTime.append(rownum)
#     return 0
#
# def Test(rownum):
#     cell1 = 'L' + str(rownum)
#     cell2 = 'M' + str(rownum)
#     Test1 = GD_sheet[cell1]
#     Test2 = GD_sheet[cell2]
#     guzhangxiaochushijian = re.match(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}',str(Test1.value)).group()
#     try:
#         gaojingqingchushijian = re.match(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}',str(Test2.value)).group()
#     except AttributeError:
#         print ('告警消除时间为空')
#         return 0
#     print (guzhangxiaochushijian, gaojingqingchushijian)
#     time1 = datetime.datetime.strptime(guzhangxiaochushijian,'%Y-%m-%d %H:%M')
#     time2 = datetime.datetime.strptime(gaojingqingchushijian,'%Y-%m-%d %H:%M')
#     if time1 < time2 :
#         print ('故障消除时间早于告警清除时间')
#     else :
#         print ('时间正确')
#
# measureTimeList = []
# def takeMeasuresTime(rownum):
#     cell1 = 'L' + str(rownum)
#     cell2 = 'F' + str(rownum)
#     cell3 = 'Y' + str(rownum)
#     timeList = [str(GD_sheet[cell1].value), str(GD_sheet[cell2].value), str(GD_sheet[cell3].value)]
#     for i in range(len(timeList)) :
#         timeList[i] = formatChange(timeList[i])
#     for i in range(len(timeList)) :
#         result = timeType(timeList[i])
#         if result == 0 :
#             print ('时间为空值,跳过')
#             return 0
#         elif result == 1:
#             timeList[i] += ':00'
#     print (timeList)
#     clearTime = datetime.datetime.strptime(timeList[0],'%Y-%m-%d %H:%M:%S')
#     happenTime = datetime.datetime.strptime(timeList[1],'%Y-%m-%d %H:%M:%S')
#     measureTime = datetime.datetime.strptime(timeList[2],'%Y-%m-%d %H:%M:%S')
#     if not (measureTime >= happenTime and measureTime <= clearTime):
#         measureTimeList.append(rownum)
#     return 0
#
# def timeType(cellvalue):
#     secondType = re.match(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}',str(cellvalue))
#     minuteType = re.match(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}',str(cellvalue))
#     emptyType =  re.match(r'\s+',str(cellvalue))
#     if emptyType :
#         return 0
#     elif  not (minuteType and secondType):
#         return 1
#
# def formatChange(cellvalue):
#     valueRe = re.split('\.', cellvalue)
#     return valueRe[0]
#
# takeMeasuresTime(14)
# print (measureTimeList)

# preProcessNotMatchList =[]
#
# def preProcessNotMatch(rownum):
#     cell1 = 'O' + str(rownum)
#     cell2 = 'R' + str(rownum)
#     cell3 = 'AJ' + str(rownum)
#     if '智能' in GD_sheet[cell3].value :
#         Reason = []
#         if '智能原因' in GD_sheet[cell3].value:
#             Reason = re.split(r'智能原因：', GD_sheet[cell3].value)
#         elif '智能定位原因' in GD_sheet[cell3].value:
#             Reason = re.split(r'智能定位原因（', GD_sheet[cell3].value)
#         if Reason[1][0:2] == '不确':
#             print ('理由不确定')
#             return 0
#         AIReason = Reason[1][0:2]
#     else :
#         print (rownum,'行预处理字段为空')
#         return 0
#     failureCauseCategory = GD_sheet[cell1].value
#     FCCRe = re.split('->', failureCauseCategory)
#     FCCReason = FCCRe[1]
#     description = GD_sheet[cell2].value
#     if AIReason != FCCReason and '预处理' not in description:
#         print ('故障原因分类与预处理结果不匹配')
#         preProcessNotMatchList.append(rownum)


# for i in range(2,12):
#     preProcessNotMatch(i)
# print (preProcessNotMatchList)


#oldExcel = input(r'请输入需要处理的工单的路径以及文件，请在文件后加上.xlsx。\n如：C:\Users\Admin\Desktop\工单.xlsx :')


def timeType(cellvalue):
    secondType = re.match(r'\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}:\d{1,2}', str(cellvalue))
    minuteType = re.match(r'\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}', str(cellvalue))
    emptyType = re.match(r'\s+', str(cellvalue))
    if emptyType:
        return 0
    elif not (minuteType and secondType):
        return 1


