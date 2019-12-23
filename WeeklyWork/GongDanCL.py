from openpyxl import load_workbook
import re
import datetime
import sys

try:
    oldExcel = input(r'请输入需要处理的工单的路径以及文件，请在文件后加上.xlsx。\n如：C:\Users\Admin\Desktop\工单.xlsx :')
    oldExcel = re.sub('"', '', oldExcel)
    newExcel = oldExcel[0:oldExcel.find('.')] + '筛选后结果.xlsx'
    print('筛选后的表格将会在 %s 路径下生成' % newExcel)
    input('输入回车将继续执行')
    GD_excel = load_workbook(r'%s' % oldExcel)
except Exception as e:
    print('文件读取失败，请尝试将文件打开后保存一次再执行程序或检查文件是否存在或文件格式是否正确。')
    print('错误为: %s' % e)
    input('输入回车结束')
    sys.exit()

GD_sheet = GD_excel['故障工单查询导出']
GD_Error = GD_excel.create_sheet('问题工单汇总')
print('工单文件加载完成，新工单表格生成中...')

# 将原表格的表头复制到新工作表中去
title = GD_sheet[1]
a = 1
for i in title:
    GD_Error.cell(row=1, column=a, value=i.value)
    a += 1
GD_Error.cell(row=1, column=GD_sheet.max_column + 1, value='工单错误原因')

recoverList, autoRecoverList, deprecationTimeList, processTimeList, typeCompareList, clearWarnCompareList, measureTimeList, preProcessNotMatchList = [], [], [], [], [], [], [], []
clearTimeError, warnClearTimeError, measureTimeError, happenTimeError = [], [], [], []


# 判断“故障原因类别“中是否包含”恢复“字样
def recoverOrNot(linenum):
    cell = 'O' + str(linenum)
    failureCauseCategory = GD_sheet[cell]
    if failureCauseCategory.value.find('恢复') != -1:
        recoverList.append(linenum)
    return 0


# 在“问题工单汇总”工作表中插入新的一行并将指定的工单信息复制，并在最后添加工单错误的原因errorinfo
def insertData(linenum, errorInfo):
    GD_Error.insert_rows(2)
    a = 1
    for i in GD_sheet[linenum]:
        GD_Error.cell(row=2, column=a, value=i.value)
        a += 1
        GD_Error.cell(row=2, column=a, value=errorInfo)
    return 0


# 判断“说明（故障处理结果）”中是否包含“自动恢复”字样，插入不包含的行号到列表autoRecoverList中
def autoRecover(linenum):
    cell = 'R' + str(linenum)
    Description = GD_sheet[cell]
    if Description.value.find('自动恢复') == -1:
        autoRecoverList.append(linenum)
    return 0


# 匹配规则是将中文的一些标点符号转换成英文的防止使用函数时匹配出现错误
table = {ord(f): ord(t) for f, t in zip(u'，。！？【】（）％＃＠＆１２３４５６７８９０－：', u',.!?[]()%#@&1234567890-:')}


# 判断“说明”中的时间是否与“故障消除时间”一致，将不一致的工单的行号插入deprecationTime中，时间精确到分钟
def troubleShootingTimeCompare(rownum):
    cell1 = 'L' + str(rownum)
    troubleShootingTime = re.match(r'\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}',
                                   str(GD_sheet[cell1].value).translate(table)).group()
    cell2 = 'R' + str(rownum)
    descriptionCell = re.findall(r'\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}', str(GD_sheet[cell2].value).translate(table))
    if descriptionCell == []:
        return 0
    if troubleShootingTime != descriptionCell[0]:
        deprecationTimeList.append(rownum)
    return 0


# 判断‘故障处理历时’是否大于等于24小时
def processTime(rownum):
    cell = 'J' + str(rownum)
    processTimeCell = str(GD_sheet[cell].value)
    if re.match('/s+', processTimeCell):
        return 0
    processTimeVaule = float(processTimeCell)
    if processTimeVaule >= 24:
        processTimeList.append(rownum)
    return 0


# 判断‘工单类别’和‘故障原因类别’的第二目录是否相符
def typeCompare(rownum):
    cell1 = 'A' + str(rownum)
    cell2 = 'O' + str(rownum)
    GDtype = GD_sheet[cell1].value
    failureCauseCategory = GD_sheet[cell2].value
    FCCre = re.split('->', failureCauseCategory)
    if GDtype not in FCCre[1]:
        typeCompareList.append(rownum)
    return 0


# 判断故障消除时间早于告警清除时间的工单
def clearAndWarnTimeCompare(rownum):
    cell1 = 'L' + str(rownum)
    cell2 = 'M' + str(rownum)
    clearTimeRe = re.sub(' ', '', str(GD_sheet[cell1].value))
    warnClearTimeRe = re.sub(' ', '', str(GD_sheet[cell2].value))
    try:
        clearTime = re.match(r'\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}', clearTimeRe).group()
        warnClearTime = re.match(r'\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}', warnClearTimeRe).group()
    except AttributeError:
        return 0
    try:
        time1 = datetime.datetime.strptime(clearTime, '%Y-%m-%d %H:%M')
    except ValueError:
        print('出现故障消除时间格式错误的工单。请勿关闭窗口，程序会继续执行。')
        clearTimeError.append(rownum)
        return 0
    try:
        time2 = datetime.datetime.strptime(warnClearTime, '%Y-%m-%d %H:%M')
    except ValueError:
        print('出现告警清除时间格式错误的工单。请勿关闭窗口，程序会继续执行。')
        warnClearTimeError.append(rownum)
        return 0
    if time1 < time2:
        clearWarnCompareList.append(rownum)
    return 0


# “采取措施时间”要大于等于“发生时间”，且小于等于“故障消除时间”，提取出不在此范围内的工单
def takeMeasuresTime(rownum):
    cell1 = 'L' + str(rownum)
    cell2 = 'F' + str(rownum)
    cell3 = 'Y' + str(rownum)
    timeList = [str(GD_sheet[cell1].value), str(GD_sheet[cell2].value), str(GD_sheet[cell3].value)]
    for i in range(len(timeList)):
        timeList[i] = formatChange(timeList[i])
    for i in range(len(timeList)):
        result = timeType(timeList[i])
        if result == 0:
            return 0
        elif result == 1:
            timeList[i] += ':00'
    try:
        clearTime = datetime.datetime.strptime(timeList[0], '%Y-%m-%d %H:%M:%S')
    except ValueError as e:
        print('出现故障消除时间格式错误的工单。请勿关闭窗口，程序会继续执行。')
        clearTimeError.append(rownum)
        return 0
    try:
        happenTime = datetime.datetime.strptime(timeList[1], '%Y-%m-%d %H:%M:%S')
    except ValueError :
        print('出现故障发生时间格式错误的工单。请勿关闭窗口，程序会继续执行。')
        happenTimeError.append(rownum)
        return 0
    try:
        measureTime = datetime.datetime.strptime(timeList[2], '%Y-%m-%d %H:%M:%S')
    except ValueError :
        print('出现采取措施时间格式错误的工单。请勿关闭窗口，程序会继续执行。')
        measureTimeError.append(rownum)
        return 0
    if not (measureTime >= happenTime and measureTime <= clearTime):
        measureTimeList.append(rownum)
    return 0


# 判断是否该单元格对应的时间是空值或者未精确到秒，对其进行分类
def timeType(cellvalue):
    secondType = re.match(r'\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}:\d{1,2}', str(cellvalue))
    minuteType = re.match(r'\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}', str(cellvalue))
    emptyType = re.match(r'\s+', str(cellvalue))
    if emptyType:
        return 0
    elif not (minuteType and secondType):
        return 1


# 修改单元格中时间的格式，有些时间会在取值后出现秒位中有小数存在的情况，将其清除
def formatChange(cellvalue):
    valueRe = re.split('\.', cellvalue)
    return valueRe[0]


# 判断‘预处理’内容中是否有‘智能’字样，取出智能原因与‘故障原因类别’的第二个字段对比，再判断‘说明’中是否含有‘预处理字样’。提取出不符合条件的工单的行号
def preProcessNotMatch(rownum):
    cell1 = 'O' + str(rownum)
    cell2 = 'R' + str(rownum)
    cell3 = 'AJ' + str(rownum)
    if '智能' in GD_sheet[cell3].value:
        Reason = []
        if '智能原因' in GD_sheet[cell3].value:
            Reason = re.split(r'智能原因：', GD_sheet[cell3].value)
        elif '智能定位原因' in GD_sheet[cell3].value:
            Reason = re.split(r'智能定位原因（', GD_sheet[cell3].value)
        if Reason[1][0:2] == '不确':
            return 0
        AIReason = Reason[1][0:2]
    else:
        return 0
    failureCauseCategory = GD_sheet[cell1].value
    FCCRe = re.split('->', failureCauseCategory)
    FCCReason = FCCRe[1]
    description = GD_sheet[cell2].value
    if AIReason not in FCCReason and '预处理' not in description:
        preProcessNotMatchList.append(rownum)


# 循环执行函数，将‘故障原因分类与预处理结果不匹配’类型的工单号提取出来保存到preProcessNotMatchList列表中
for num in range(2, GD_sheet.max_row):
    preProcessNotMatch(num)

for num in preProcessNotMatchList:
    insertData(num, '故障原因分类与预处理结果不匹配')

# 循环执行函数，“采取措施时间”要大于等于“发生时间”，且小于等于“故障消除时间”，提取出不在此范围内的工单，将行号写入measureTimeList中
for num in range(2, GD_sheet.max_row):
    takeMeasuresTime(num)

for num in measureTimeList:
    insertData(num, '采取措施时间填写不合理')

# 循环执行函数，匹配故障时间早于告警清除时间的工单，保存行号到clearWarnCompareList中
for num in range(2, GD_sheet.max_row):
    clearAndWarnTimeCompare(num)

for num in clearWarnCompareList:
    insertData(num, '故障消除时间早于告警清除时间')

# 循环执行函数，匹配故障原因类别中包含恢复字样的工单，保留行号到列表recoverList中
for num in range(2, GD_sheet.max_row):
    recoverOrNot(num)

# 循环执行函数，判断‘工单类别’和‘故障原因类别’中的第二目录是否存在包含关系
for num in recoverList:
    typeCompare(num)

for num in typeCompareList:
    insertData(num, '跨专业报结自动恢复')

# 循环执行函数，在'故障原因类别'含有‘恢复’字样的工单中筛选出‘故障处理历时’大于等于24的工单，并插入到新工作表中
for num in recoverList:
    processTime(num)

for num in processTimeList:
    insertData(num, '故障历时超24小时报结自动恢复')

# 循环执行函数，在'故障原因类别'含有‘恢复’字样的工单中筛选是否有工单的‘说明’中的时间与‘故障消除时间’是否存在差异
for num in recoverList:
    troubleShootingTimeCompare(num)

for num in deprecationTimeList:
    insertData(num, '说明内时间填写不正确')

# 循环执行函数，在'故障原因类别'含有‘恢复’字样的工单中筛选是否有工单的‘说明’中不包含‘自动恢复’字样
for num in recoverList:
    autoRecover(num)

for num in autoRecoverList:
    insertData(num, '自动恢复报结不规范')

# 将时间格式不正确的工单写入新工作表
for num in clearTimeError:
    insertData(num, '故障消除时间格式填写错误')
for num in warnClearTimeError:
    insertData(num, '告警消除时间格式填写错误')
for num in happenTimeError:
    insertData(num, '故障发生时间格式填写错误')
for num in measureTimeError:
    insertData(num, '采取措施时间格式填写错误')

try:
    GD_excel.save(r'%s' % newExcel)
except PermissionError:
    print('权限错误，无法保存为目标文件。请检查目标文件是否已经被打开从而无权限对其进行读写。')
except Exception as e:
    print('生成文件失败,错误为：%s' % e)
print(r'目标文件 %s 已生成。' % newExcel)
input("输入回车结束")
